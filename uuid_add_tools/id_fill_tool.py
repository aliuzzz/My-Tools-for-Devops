# -*- coding: utf-8 -*-
"""
识别XLS/XLSX的ID列，填充空缺的ID列（唯一UUID），并生成新的XLS/XLSX文件。
"""

from __future__ import annotations

import csv
import re
import secrets
import shutil
import threading
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from queue import Empty, Queue
from typing import Callable, Iterable

import tkinter as tk
from tkinter import filedialog, ttk

import xlrd
import xlwt
from openpyxl import Workbook, load_workbook

# 尝试导入 tkinterdnd2 以支持拖拽功能
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD

    HAS_DND = True
except ImportError:
    HAS_DND = False


ALLOWED_ID_CHARS = "23456789ABCDEFGHJKMNPQRSTUVWXYZ"
ID_GROUP_LENGTHS = (4, 4, 5)
INPUT_SUFFIXES = (".xlsx", ".xls")
OUTPUT_SUFFIXES = {
    "csv": ".csv",
    "xls": ".xls",
    "xlsx": ".xlsx",
}
YEAR_PATTERN = re.compile(r"(?<!\d)(\d{4})(?!\d)")
EXPORT_CACHE: dict[tuple[str, str, int, int], "PreparedExport"] = {}
ProgressCallback = Callable[[int, str], None]


class ProcessingError(Exception):
    """Raised when the selected file cannot be processed safely."""


@dataclass
class ProcessResult:
    output_path: Path
    added_count: int
    content_row_count: int
    preserved_id_count: int
    filtered_row_count: int | None = None


@dataclass
class PreparedSheet:
    rows: list[list[object]]
    id_index: int
    added_count: int
    content_row_count: int
    preserved_id_count: int

    def clone(self) -> "PreparedSheet":
        return PreparedSheet(
            rows=clone_rows(self.rows),
            id_index=self.id_index,
            added_count=self.added_count,
            content_row_count=self.content_row_count,
            preserved_id_count=self.preserved_id_count,
        )


@dataclass
class WorkbookData:
    sheet_names: list[str]
    sheets: dict[str, list[list[object]]]

    def clone(self) -> "WorkbookData":
        return WorkbookData(
            sheet_names=self.sheet_names.copy(),
            sheets={sheet_name: clone_rows(self.sheets[sheet_name]) for sheet_name in self.sheet_names},
        )


@dataclass
class PreparedExport:
    workbook: WorkbookData
    selected_sheet: str
    prepared_sheet: PreparedSheet

    def clone(self) -> "PreparedExport":
        return PreparedExport(
            workbook=self.workbook.clone(),
            selected_sheet=self.selected_sheet,
            prepared_sheet=self.prepared_sheet.clone(),
        )


def clone_rows(rows: list[list[object]]) -> list[list[object]]:
    return [row.copy() for row in rows]


def report_progress(progress_callback: ProgressCallback | None, percent: int, message: str) -> None:
    if progress_callback is None:
        return
    progress_callback(max(0, min(100, percent)), message)


def report_progress_in_range(
    progress_callback: ProgressCallback | None,
    start_percent: int,
    end_percent: int,
    current_index: int,
    total_count: int,
    message: str,
) -> None:
    if progress_callback is None or total_count <= 0:
        return
    ratio = current_index / total_count
    percent = start_percent + int((end_percent - start_percent) * ratio)
    report_progress(progress_callback, percent, message)


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_id(value: object) -> str:
    return normalize_cell(value).upper()


def find_column(headers: Iterable[object], target: str, *, case_sensitive: bool) -> int:
    expected = target if case_sensitive else target.upper()
    for index, value in enumerate(headers):
        current = normalize_cell(value)
        if not case_sensitive:
            current = current.upper()
        if current == expected:
            return index
    raise ProcessingError(f"首行没有找到名为 {target} 的列。")


def find_id_column(headers: Iterable[object]) -> int:
    return find_column(headers, "ID", case_sensitive=True)


def find_period_column(headers: Iterable[object]) -> int:
    return find_column(headers, "对应区间", case_sensitive=True)


def row_has_content(row_values: list[object], id_index: int) -> bool:
    return any(normalize_cell(value) for index, value in enumerate(row_values) if index != id_index)


def generate_unique_id(existing_ids: set[str]) -> str:
    while True:
        candidate = "-".join(
            "".join(secrets.choice(ALLOWED_ID_CHARS) for _ in range(group_length))
            for group_length in ID_GROUP_LENGTHS
        )
        if candidate not in existing_ids:
            existing_ids.add(candidate)
            return candidate


def build_output_path(input_path: Path, output_suffix: str, label: str | None = None) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{input_path.stem}_{timestamp}" if not label else f"{input_path.stem}_{label}_{timestamp}"
    candidate = input_path.with_name(f"{base_name}{output_suffix}")
    counter = 1
    while candidate.exists():
        candidate = input_path.with_name(f"{base_name}_{counter}{output_suffix}")
        counter += 1
    return candidate


def validate_existing_ids(
    rows: list[list[object]],
    id_index: int,
    progress_callback: ProgressCallback | None = None,
) -> tuple[set[str], list[str], int]:
    seen: dict[str, int] = {}
    duplicates: list[str] = []
    content_row_count = 0

    total_rows = max(len(rows) - 1, 1)
    for offset, (row_number, row_values) in enumerate(enumerate(rows[1:], start=2), start=1):
        current_id = normalize_id(row_values[id_index]) if id_index < len(row_values) else ""
        if row_has_content(row_values, id_index):
            content_row_count += 1
        if not current_id:
            report_progress_in_range(progress_callback, 16, 40, offset, total_rows, "正在检查现有 ID...")
            continue
        if current_id in seen:
            duplicates.append(f"第 {seen[current_id]} 行 和 第 {row_number} 行: {current_id}")
            report_progress_in_range(progress_callback, 16, 40, offset, total_rows, "正在检查现有 ID...")
            continue
        seen[current_id] = row_number
        report_progress_in_range(progress_callback, 16, 40, offset, total_rows, "正在检查现有 ID...")

    return set(seen), duplicates, content_row_count


def prepare_sheet(rows: list[list[object]], progress_callback: ProgressCallback | None = None) -> PreparedSheet:
    if not rows:
        raise ProcessingError("表格内容为空。")

    report_progress(progress_callback, 10, "正在定位 ID 列...")
    id_index = find_id_column(rows[0])
    existing_ids, duplicates, content_row_count = validate_existing_ids(rows, id_index, progress_callback)
    if duplicates:
        duplicate_text = "\n".join(duplicates[:10])
        if len(duplicates) > 10:
            duplicate_text += f"\n... 另外还有 {len(duplicates) - 10} 处重复"
        raise ProcessingError(f"发现现有 ID 重复，已停止生成：\n{duplicate_text}")

    report_progress(progress_callback, 40, "正在补充缺失 ID...")
    added_count = 0
    total_rows = max(len(rows) - 1, 1)
    for offset, row_values in enumerate(rows[1:], start=1):
        missing_cells = id_index - len(row_values) + 1
        if missing_cells > 0:
            row_values.extend([""] * missing_cells)
        if row_has_content(row_values, id_index) and not normalize_id(row_values[id_index]):
            row_values[id_index] = generate_unique_id(existing_ids)
            added_count += 1
        report_progress_in_range(progress_callback, 40, 70, offset, total_rows, "正在补充缺失 ID...")

    if added_count == 0:
        raise ProcessingError("没有找到需要补 ID 的有效行，原文件未修改。")

    report_progress(progress_callback, 70, "ID 补充完成，正在准备导出...")
    return PreparedSheet(rows, id_index, added_count, content_row_count, len(existing_ids) - added_count)


def cell_matches_year(value: object, target_year: int) -> bool:
    if isinstance(value, datetime):
        return value.year == target_year
    if isinstance(value, date):
        return value.year == target_year

    text = normalize_cell(value)
    if not text:
        return False
    return any(int(year_text) == target_year for year_text in YEAR_PATTERN.findall(text))


def filter_rows_by_year(
    rows: list[list[object]],
    target_year: int,
    progress_callback: ProgressCallback | None = None,
) -> tuple[list[list[object]], int, list[int]]:
    period_index = find_period_column(rows[0])
    filtered_rows = [rows[0]]
    kept_source_row_numbers: list[int] = []

    total_rows = max(len(rows) - 1, 1)
    for offset, (row_number, row_values) in enumerate(enumerate(rows[1:], start=2), start=1):
        if period_index < len(row_values) and cell_matches_year(row_values[period_index], target_year):
            filtered_rows.append(row_values.copy())
            kept_source_row_numbers.append(row_number)
        report_progress_in_range(progress_callback, 70, 82, offset, total_rows, "正在筛选本年度数据...")

    filtered_row_count = len(filtered_rows) - 1
    if filtered_row_count <= 0:
        raise ProcessingError(f"没有找到“对应区间”为 {target_year} 年的行，未导出年度副本。")

    return filtered_rows, filtered_row_count, kept_source_row_numbers


def build_file_signature(file_path: Path) -> tuple[str, int, int]:
    resolved_path = file_path.resolve()
    stat_result = resolved_path.stat()
    return str(resolved_path), stat_result.st_mtime_ns, stat_result.st_size


def build_cache_key(file_path: Path, sheet_name: str) -> tuple[str, str, int, int]:
    path_text, modified_ns, file_size = build_file_signature(file_path)
    return path_text, sheet_name, modified_ns, file_size


def get_sheet_names(file_path: Path) -> list[str]:
    file_suffix = file_path.suffix.lower()
    if file_suffix == ".xlsx":
        workbook = load_workbook(file_path, read_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()
    if file_suffix == ".xls":
        workbook = xlrd.open_workbook(file_path)
        return workbook.sheet_names()
    raise ProcessingError("仅支持 xlsx 和 xls 文件。")


def read_workbook_data(file_path: Path) -> WorkbookData:
    file_suffix = file_path.suffix.lower()
    if file_suffix == ".xlsx":
        return read_xlsx_workbook(file_path)
    if file_suffix == ".xls":
        return read_xls_workbook(file_path)
    raise ProcessingError("仅支持 xlsx 和 xls 文件。")


def read_xlsx_workbook(file_path: Path, progress_callback: ProgressCallback | None = None) -> WorkbookData:
    workbook = load_workbook(file_path, data_only=False)
    try:
        report_progress(progress_callback, 4, "正在读取工作簿...")
        total_sheets = max(len(workbook.sheetnames), 1)
        sheets = {
            sheet_name: read_xlsx_sheet(workbook[sheet_name], progress_callback, index, total_sheets)
            for index, sheet_name in enumerate(workbook.sheetnames, start=1)
        }
        return WorkbookData(list(workbook.sheetnames), sheets)
    finally:
        workbook.close()


def read_xlsx_sheet(worksheet, progress_callback: ProgressCallback | None, sheet_index: int, total_sheets: int) -> list[list[object]]:
    report_progress_in_range(progress_callback, 4, 10, sheet_index, total_sheets, f"正在读取工作表：{worksheet.title}")
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def read_xls_workbook(file_path: Path, progress_callback: ProgressCallback | None = None) -> WorkbookData:
    workbook = xlrd.open_workbook(file_path)
    sheets: dict[str, list[list[object]]] = {}
    sheet_names = workbook.sheet_names()

    total_sheets = max(len(sheet_names), 1)
    for index, sheet_name in enumerate(sheet_names, start=1):
        worksheet = workbook.sheet_by_name(sheet_name)
        rows = [
            [read_xls_cell(workbook, worksheet, row_index, column_index) for column_index in range(worksheet.ncols)]
            for row_index in range(worksheet.nrows)
        ]
        sheets[sheet_name] = rows
        report_progress_in_range(progress_callback, 4, 10, index, total_sheets, f"正在读取工作表：{sheet_name}")

    return WorkbookData(sheet_names, sheets)


def read_xls_cell(workbook: xlrd.book.Book, worksheet: xlrd.sheet.Sheet, row_index: int, column_index: int) -> object:
    cell = worksheet.cell(row_index, column_index)
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return int(cell.value) if float(cell.value).is_integer() else cell.value
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
        return ""
    return cell.value


def prepare_workbook_export(
    file_path: Path,
    sheet_name: str,
    progress_callback: ProgressCallback | None = None,
) -> PreparedExport:
    cache_key = build_cache_key(file_path, sheet_name)
    cached_export = EXPORT_CACHE.get(cache_key)
    if cached_export is not None:
        report_progress(progress_callback, 70, "已复用缓存结果，正在准备导出...")
        return cached_export.clone()

    workbook = read_workbook_data_with_progress(file_path, progress_callback)
    if sheet_name not in workbook.sheets:
        raise ProcessingError(f"没有找到工作表: {sheet_name}")

    prepared_sheet = prepare_sheet(clone_rows(workbook.sheets[sheet_name]), progress_callback)
    workbook.sheets[sheet_name] = clone_rows(prepared_sheet.rows)

    prepared_export = PreparedExport(workbook, sheet_name, prepared_sheet)
    EXPORT_CACHE[cache_key] = prepared_export
    return prepared_export.clone()


def read_workbook_data_with_progress(file_path: Path, progress_callback: ProgressCallback | None = None) -> WorkbookData:
    file_suffix = file_path.suffix.lower()
    if file_suffix == ".xlsx":
        return read_xlsx_workbook(file_path, progress_callback)
    if file_suffix == ".xls":
        return read_xls_workbook(file_path, progress_callback)
    raise ProcessingError("仅支持 xlsx 和 xls 文件。")


def export_file(
    file_path: Path,
    sheet_name: str,
    output_format: str,
    current_year_only: bool,
    progress_callback: ProgressCallback | None = None,
) -> ProcessResult:
    if output_format not in OUTPUT_SUFFIXES:
        raise ProcessingError("导出格式仅支持 csv、xls 和 xlsx。")

    report_progress(progress_callback, 0, "开始处理...")
    prepared_export = prepare_workbook_export(file_path, sheet_name, progress_callback)
    workbook = prepared_export.workbook
    filtered_row_count: int | None = None
    kept_source_row_numbers: list[int] | None = None
    label: str | None = None

    if current_year_only:
        target_year = datetime.now().year
        filtered_rows, filtered_row_count, kept_source_row_numbers = filter_rows_by_year(
            workbook.sheets[sheet_name], target_year, progress_callback
        )
        workbook.sheets[sheet_name] = filtered_rows
        label = f"{target_year}年度"
    else:
        report_progress(progress_callback, 82, "正在生成导出文件...")

    output_path = build_output_path(file_path, OUTPUT_SUFFIXES[output_format], label)
    write_output_file(
        output_path=output_path,
        workbook=workbook,
        selected_sheet=sheet_name,
        output_format=output_format,
        source_path=file_path,
        prepared_sheet=prepared_export.prepared_sheet,
        kept_source_row_numbers=kept_source_row_numbers,
        progress_callback=progress_callback,
    )
    report_progress(progress_callback, 100, "处理完成")

    return ProcessResult(
        output_path=output_path,
        added_count=prepared_export.prepared_sheet.added_count,
        content_row_count=prepared_export.prepared_sheet.content_row_count,
        preserved_id_count=prepared_export.prepared_sheet.preserved_id_count,
        filtered_row_count=filtered_row_count,
    )


def write_output_file(
    output_path: Path,
    workbook: WorkbookData,
    selected_sheet: str,
    output_format: str,
    source_path: Path,
    prepared_sheet: PreparedSheet,
    kept_source_row_numbers: list[int] | None,
    progress_callback: ProgressCallback | None,
) -> None:
    if output_format == "csv":
        write_csv_file(output_path, workbook.sheets[selected_sheet], progress_callback)
        return
    if output_format == "xlsx":
        if source_path.suffix.lower() == ".xlsx":
            write_xlsx_file_preserving_source(
                output_path=output_path,
                source_path=source_path,
                selected_sheet=selected_sheet,
                prepared_sheet=prepared_sheet,
                kept_source_row_numbers=kept_source_row_numbers,
                progress_callback=progress_callback,
            )
            return
        write_xlsx_file(output_path, workbook, progress_callback)
        return
    if output_format == "xls":
        write_xls_file(output_path, workbook, progress_callback)
        return
    raise ProcessingError("导出格式仅支持 csv、xls 和 xlsx。")


def write_csv_file(output_path: Path, rows: list[list[object]], progress_callback: ProgressCallback | None = None) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        total_rows = max(len(rows), 1)
        for index, row_values in enumerate(rows, start=1):
            writer.writerow(row_values)
            report_progress_in_range(progress_callback, 82, 98, index, total_rows, "正在写出 CSV 文件...")


def write_xlsx_file(
    output_path: Path,
    workbook_data: WorkbookData,
    progress_callback: ProgressCallback | None = None,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    total_rows = max(sum(len(workbook_data.sheets[sheet_name]) for sheet_name in workbook_data.sheet_names), 1)
    written_rows = 0
    for sheet_name in workbook_data.sheet_names:
        worksheet = workbook.create_sheet(title=sheet_name)
        for row_values in workbook_data.sheets[sheet_name]:
            worksheet.append([normalize_xlsx_value(value) for value in row_values])
            written_rows += 1
            report_progress_in_range(progress_callback, 82, 96, written_rows, total_rows, "正在写出 XLSX 文件...")

    report_progress(progress_callback, 98, "正在保存 XLSX 文件...")
    workbook.save(output_path)
    workbook.close()


def write_xlsx_file_preserving_source(
    output_path: Path,
    source_path: Path,
    selected_sheet: str,
    prepared_sheet: PreparedSheet,
    kept_source_row_numbers: list[int] | None,
    progress_callback: ProgressCallback | None = None,
) -> None:
    report_progress(progress_callback, 82, "正在复制原始工作簿...")
    shutil.copy2(source_path, output_path)
    report_progress(progress_callback, 84, "正在打开复制后的工作簿...")
    workbook = load_workbook(output_path)
    try:
        worksheet = workbook[selected_sheet]
        write_id_column_to_worksheet(worksheet, prepared_sheet, progress_callback)

        if kept_source_row_numbers is not None:
            keep_row_set = set(kept_source_row_numbers)
            total_rows = max(worksheet.max_row - 1, 1)
            processed_rows = 0
            for row_number in range(worksheet.max_row, 1, -1):
                if row_number not in keep_row_set:
                    worksheet.delete_rows(row_number)
                processed_rows += 1
                report_progress_in_range(progress_callback, 92, 98, processed_rows, total_rows, "正在筛除非本年度行...")

        report_progress(progress_callback, 98, "正在保存 XLSX 文件...")
        workbook.save(output_path)
    finally:
        workbook.close()


def write_id_column_to_worksheet(
    worksheet,
    prepared_sheet: PreparedSheet,
    progress_callback: ProgressCallback | None = None,
) -> None:
    total_rows = max(len(prepared_sheet.rows) - 1, 1)
    for offset, (row_number, row_values) in enumerate(enumerate(prepared_sheet.rows[1:], start=2), start=1):
        worksheet.cell(row=row_number, column=prepared_sheet.id_index + 1, value=row_values[prepared_sheet.id_index])
        report_progress_in_range(progress_callback, 84, 92, offset, total_rows, "正在写入 ID 到工作簿...")


def normalize_xlsx_value(value: object) -> object:
    if isinstance(value, date) and not isinstance(value, datetime):
        return datetime(value.year, value.month, value.day)
    return value


def write_xls_file(
    output_path: Path,
    workbook_data: WorkbookData,
    progress_callback: ProgressCallback | None = None,
) -> None:
    workbook = xlwt.Workbook(encoding="utf-8")
    date_style = xlwt.easyxf(num_format_str="YYYY-MM-DD")
    datetime_style = xlwt.easyxf(num_format_str="YYYY-MM-DD HH:MM:SS")

    total_rows = max(sum(len(workbook_data.sheets[sheet_name]) for sheet_name in workbook_data.sheet_names), 1)
    written_rows = 0
    for sheet_name in workbook_data.sheet_names:
        worksheet = workbook.add_sheet(sheet_name)
        for row_index, row_values in enumerate(workbook_data.sheets[sheet_name]):
            for column_index, value in enumerate(row_values):
                write_xls_cell(worksheet, row_index, column_index, value, date_style, datetime_style)
            written_rows += 1
            report_progress_in_range(progress_callback, 82, 98, written_rows, total_rows, "正在写出 XLS 文件...")

    workbook.save(str(output_path))


def write_xls_cell(
    worksheet: xlwt.Worksheet,
    row_index: int,
    column_index: int,
    value: object,
    date_style: xlwt.XFStyle,
    datetime_style: xlwt.XFStyle,
) -> None:
    if value is None or value == "":
        worksheet.write(row_index, column_index, "")
        return
    if isinstance(value, datetime):
        worksheet.write(row_index, column_index, value, datetime_style)
        return
    if isinstance(value, date):
        worksheet.write(row_index, column_index, datetime(value.year, value.month, value.day), date_style)
        return
    worksheet.write(row_index, column_index, value)


class IdFillApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("ID 补全工具")
        self.root.resizable(False, False)
        self.root.geometry("760x540")
        self.root.configure(bg="#f3f6fb")

        self.file_path_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.output_format_var = tk.StringVar(value="xlsx")
        self.status_var = tk.StringVar(value=self._build_idle_status())
        self.progress_text_var = tk.StringVar(value="")

        self.sheet_combo: ttk.Combobox | None = None
        self.pick_button: ttk.Button | None = None
        self.export_button: ttk.Button | None = None
        self.year_export_button: ttk.Button | None = None
        self.progress_bar: ttk.Progressbar | None = None
        self.progress_label: ttk.Label | None = None
        self.format_buttons: list[ttk.Radiobutton] = []
        self.processing = False
        self.result_queue: Queue[tuple[str, object]] = Queue()
        self._configure_styles()
        self.build_ui()
        self._setup_dnd()

    def _build_idle_status(self) -> str:
        drop_hint = "也可以直接把文件拖进窗口。" if HAS_DND else "当前界面保留文件选择方式。"
        return f"请选择一个 xlsx 或 xls 文件。{drop_hint}"

    def _configure_styles(self) -> None:
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        app_bg = "#f3f6fb"
        card_bg = "#ffffff"
        accent = "#1f6feb"
        accent_active = "#1758b8"
        accent_soft = "#e8f0ff"
        text_main = "#162033"
        text_muted = "#5b667a"
        border = "#d8e1ee"
        field_bg = "#f8fafc"
        success_bg = "#eef5ff"
        success_border = "#d7e6ff"

        style.configure("App.TFrame", background=app_bg)
        style.configure("Card.TFrame", background=card_bg, relief="flat")
        style.configure("Status.TFrame", background=success_bg, relief="flat")
        style.configure("ProgressRow.TFrame", background=success_bg)
        style.configure(
            "Progress.Horizontal.TProgressbar",
            troughcolor="#dbe6f7",
            background=accent_active,
            darkcolor=accent_active,
            lightcolor=accent_active,
            bordercolor=success_border,
            thickness=12,
            relief="flat",
        )

        style.configure(
            "HeroTitle.TLabel",
            background=app_bg,
            foreground=text_main,
            font=("Microsoft YaHei UI", 18, "bold"),
        )
        style.configure(
            "HeroSub.TLabel",
            background=app_bg,
            foreground=text_muted,
            font=("Microsoft YaHei UI", 10),
        )
        style.configure(
            "SectionTitle.TLabel",
            background=card_bg,
            foreground=text_main,
            font=("Microsoft YaHei UI", 11, "bold"),
        )
        style.configure(
            "FieldLabel.TLabel",
            background=card_bg,
            foreground=text_muted,
            font=("Microsoft YaHei UI", 10),
        )
        style.configure(
            "Hint.TLabel",
            background=card_bg,
            foreground=text_muted,
            font=("Microsoft YaHei UI", 9),
        )
        style.configure(
            "StatusTitle.TLabel",
            background=success_bg,
            foreground=accent_active,
            font=("Microsoft YaHei UI", 10, "bold"),
        )
        style.configure(
            "StatusBody.TLabel",
            background=success_bg,
            foreground=text_main,
            font=("Microsoft YaHei UI", 10),
        )
        style.configure(
            "ProgressMeta.TLabel",
            background=success_bg,
            foreground=accent_active,
            font=("Microsoft YaHei UI", 10, "bold"),
        )
        style.configure(
            "Field.TEntry",
            fieldbackground=field_bg,
            background=field_bg,
            foreground=text_main,
            bordercolor=border,
            lightcolor=border,
            darkcolor=border,
            padding=(10, 8),
        )
        style.map(
            "Field.TEntry",
            fieldbackground=[("readonly", field_bg)],
            foreground=[("readonly", text_main)],
        )
        style.configure(
            "Field.TCombobox",
            fieldbackground=field_bg,
            background=field_bg,
            foreground=text_main,
            bordercolor=border,
            lightcolor=border,
            darkcolor=border,
            arrowsize=16,
            padding=(8, 6),
        )
        style.map(
            "Field.TCombobox",
            fieldbackground=[("readonly", field_bg), ("disabled", "#eef2f7")],
            foreground=[("readonly", text_main), ("disabled", "#8c96a8")],
        )
        style.configure(
            "Format.TRadiobutton",
            background=card_bg,
            foreground=text_main,
            font=("Microsoft YaHei UI", 10),
        )
        style.map(
            "Format.TRadiobutton",
            background=[("active", card_bg)],
            foreground=[("active", accent_active)],
        )
        style.configure(
            "Primary.TButton",
            background=accent,
            foreground="#ffffff",
            borderwidth=0,
            focusthickness=0,
            focuscolor=accent,
            padding=(14, 10),
            font=("Microsoft YaHei UI", 10, "bold"),
        )
        style.map(
            "Primary.TButton",
            background=[("active", accent_active), ("pressed", accent_active)],
            foreground=[("disabled", "#dbe7ff")],
        )
        style.configure(
            "Secondary.TButton",
            background=accent_soft,
            foreground=accent_active,
            borderwidth=0,
            focusthickness=0,
            focuscolor=accent_soft,
            padding=(14, 10),
            font=("Microsoft YaHei UI", 10, "bold"),
        )
        style.map(
            "Secondary.TButton",
            background=[("active", "#dbe8ff"), ("pressed", "#dbe8ff")],
            foreground=[("disabled", "#8ea6d6")],
        )

    def _setup_dnd(self) -> None:
        """设置拖拽支持（如果 tkinterdnd2 可用）"""
        if not HAS_DND:
            return

        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind("<<Drop>>", self._on_drop)

    def _on_drop(self, event: tk.Event) -> None:
        """处理文件拖拽事件"""
        raw_data = event.data.strip()
        file_paths = self._parse_drop_files(raw_data)
        if not file_paths:
            return
        self._apply_selected_file(Path(file_paths[0]))

    def _parse_drop_files(self, raw_data: str) -> list[str]:
        files: list[str] = []
        current = ""
        in_braces = False

        for char in raw_data:
            if char == "{":
                in_braces = True
                continue
            if char == "}":
                in_braces = False
                continue
            if char == " " and not in_braces:
                if current:
                    files.append(current)
                    current = ""
                continue
            current += char

        if current:
            files.append(current)
        return files

    def build_ui(self) -> None:
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        frame = ttk.Frame(self.root, style="App.TFrame", padding=(24, 22, 24, 22))
        frame.grid(sticky="nsew")
        frame.columnconfigure(0, weight=1)

        header = ttk.Frame(frame, style="App.TFrame")
        header.grid(row=0, column=0, sticky="ew", pady=(0, 18))
        header.columnconfigure(0, weight=1)

        ttk.Label(header, text="ID 补全工具", style="HeroTitle.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(
            header,
            text="读取 xlsx / xls，补全选中工作表的 ID，再按所选格式导出副本。",
            style="HeroSub.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(6, 0))

        file_card = ttk.Frame(frame, style="Card.TFrame", padding=(18, 18, 18, 16))
        file_card.grid(row=1, column=0, sticky="ew")
        file_card.columnconfigure(1, weight=1)

        ttk.Label(file_card, text="文件设置", style="SectionTitle.TLabel").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 14)
        )
        ttk.Label(file_card, text="文件路径", style="FieldLabel.TLabel").grid(
            row=1, column=0, sticky="w", padx=(0, 10), pady=(0, 10)
        )
        ttk.Entry(file_card, textvariable=self.file_path_var, width=58, state="readonly", style="Field.TEntry").grid(
            row=1, column=1, sticky="ew", pady=(0, 10)
        )
        self.pick_button = ttk.Button(file_card, text="选择文件", command=self.pick_file, style="Secondary.TButton")
        self.pick_button.grid(row=1, column=2, padx=(10, 0), pady=(0, 10))

        ttk.Label(file_card, text="工作表", style="FieldLabel.TLabel").grid(
            row=2, column=0, sticky="w", padx=(0, 10), pady=(0, 10)
        )
        self.sheet_combo = ttk.Combobox(
            file_card, textvariable=self.sheet_var, width=28, state="disabled", style="Field.TCombobox"
        )
        self.sheet_combo.grid(row=2, column=1, sticky="w", pady=(0, 10))

        ttk.Label(file_card, text="导出格式", style="FieldLabel.TLabel").grid(
            row=3, column=0, sticky="nw", padx=(0, 10)
        )
        format_frame = ttk.Frame(file_card, style="Card.TFrame")
        format_frame.grid(row=3, column=1, columnspan=2, sticky="w")

        for column_index, (label, value) in enumerate((("CSV", "csv"), ("XLS", "xls"), ("XLSX", "xlsx"))):
            button = ttk.Radiobutton(
                format_frame,
                text=label,
                value=value,
                variable=self.output_format_var,
                style="Format.TRadiobutton",
            )
            button.grid(row=0, column=column_index, sticky="w", padx=(0, 18))
            self.format_buttons.append(button)

        hint_text = (
            "普通副本保留全部有效行；本年度副本会在补全 ID 后，按“对应区间”筛选当前年份。"
        )
        if HAS_DND:
            hint_text += " 也可以直接把文件拖进窗口。"
        ttk.Label(file_card, text=hint_text, style="Hint.TLabel", wraplength=660, justify="left").grid(
            row=4, column=0, columnspan=3, sticky="w", pady=(14, 0)
        )

        action_row = ttk.Frame(frame, style="App.TFrame")
        action_row.grid(row=2, column=0, sticky="ew", pady=(16, 16))
        action_row.columnconfigure(0, weight=1)
        action_row.columnconfigure(1, weight=1)

        self.export_button = ttk.Button(action_row, text="导出副本", command=self.run_processing, style="Primary.TButton")
        self.export_button.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self.year_export_button = ttk.Button(
            action_row,
            text="导出本年度副本",
            command=self.run_current_year_processing,
            style="Secondary.TButton",
        )
        self.year_export_button.grid(row=0, column=1, sticky="ew", padx=(8, 0))

        status_card = ttk.Frame(frame, style="Status.TFrame", padding=(18, 16, 18, 16))
        status_card.grid(row=3, column=0, sticky="ew")
        status_card.columnconfigure(0, weight=1)

        ttk.Label(status_card, text="当前状态", style="StatusTitle.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(
            status_card,
            textvariable=self.status_var,
            style="StatusBody.TLabel",
            wraplength=660,
            justify="left",
        ).grid(row=1, column=0, sticky="w", pady=(8, 0))

        progress_row = ttk.Frame(status_card, style="ProgressRow.TFrame")
        progress_row.grid(row=2, column=0, sticky="ew", pady=(14, 0))
        progress_row.columnconfigure(0, weight=1)

        self.progress_bar = ttk.Progressbar(
            progress_row,
            mode="determinate",
            maximum=100,
            style="Progress.Horizontal.TProgressbar",
        )
        self.progress_bar.grid(row=0, column=0, sticky="ew")

        self.progress_label = ttk.Label(
            progress_row,
            textvariable=self.progress_text_var,
            style="ProgressMeta.TLabel",
            width=5,
            anchor="e",
        )
        self.progress_label.grid(row=0, column=1, sticky="e", padx=(12, 0))
        progress_row.grid_remove()

    def pick_file(self) -> None:
        selected = filedialog.askopenfilename(
            title="选择表格文件",
            filetypes=[("Excel", "*.xlsx *.xls"), ("XLSX", "*.xlsx"), ("XLS", "*.xls")],
        )
        if selected:
            self._apply_selected_file(Path(selected))

    def _set_sheet_selector(self, values: list[str], state: str, selected_value: str) -> None:
        assert self.sheet_combo is not None
        self.sheet_combo.configure(values=values, state=state)
        self.sheet_var.set(selected_value)

    def _apply_selected_file(self, file_path: Path) -> None:
        if self.processing:
            return

        file_suffix = file_path.suffix.lower()
        if file_suffix not in INPUT_SUFFIXES:
            self.status_var.set("仅支持读取 xlsx 和 xls 文件。")
            return

        self.file_path_var.set(str(file_path))
        self.output_format_var.set(file_suffix.lstrip("."))

        try:
            sheet_names = get_sheet_names(file_path)
        except Exception as exc:
            self._set_sheet_selector([], "disabled", "")
            self.status_var.set(f"读取工作表失败: {exc}")
            return

        default_sheet = sheet_names[0] if sheet_names else ""
        self._set_sheet_selector(sheet_names, "readonly", default_sheet)

        if sheet_names:
            self.status_var.set(
                f"已选择文件：{file_path.name}\n"
                f"已加载 {len(sheet_names)} 个工作表，当前默认工作表：{default_sheet}\n"
                f"当前导出格式：{self.output_format_var.get().upper()}"
            )
            return

        self.status_var.set("文件中没有可用工作表。")

    def run_processing(self) -> None:
        self._run_export(current_year_only=False)

    def run_current_year_processing(self) -> None:
        self._run_export(current_year_only=True)

    def _run_export(self, current_year_only: bool) -> None:
        if self.processing:
            return

        file_text = self.file_path_var.get().strip()
        if not file_text:
            self.status_var.set("请先选择文件。")
            return

        sheet_name = self.sheet_var.get().strip()
        if not sheet_name:
            self.status_var.set("请选择一个工作表。")
            return

        self._set_processing_state(True, "正在准备处理...")
        threading.Thread(
            target=self._export_worker,
            args=(Path(file_text), sheet_name, self.output_format_var.get().strip().lower(), current_year_only),
            daemon=True,
        ).start()
        self.root.after(120, self._poll_export_result)

    def _export_worker(self, file_path: Path, sheet_name: str, output_format: str, current_year_only: bool) -> None:
        try:
            result = export_file(
                file_path=file_path,
                sheet_name=sheet_name,
                output_format=output_format,
                current_year_only=current_year_only,
                progress_callback=self._push_progress_update,
            )
        except ProcessingError as exc:
            self.result_queue.put(("processing_error", str(exc)))
            return
        except Exception as exc:
            self.result_queue.put(("unexpected_error", str(exc)))
            return

        self.result_queue.put(("success", (result, current_year_only)))

    def _push_progress_update(self, percent: int, message: str) -> None:
        self.result_queue.put(("progress", (percent, message)))

    def _poll_export_result(self) -> None:
        final_event: tuple[str, object] | None = None
        try:
            while True:
                status, payload = self.result_queue.get_nowait()
                if status == "progress":
                    percent, message = payload
                    self._update_progress(percent, message)
                    continue
                final_event = (status, payload)
                break
        except Empty:
            if self.processing:
                self.root.after(120, self._poll_export_result)
            return

        if final_event is None:
            if self.processing:
                self.root.after(120, self._poll_export_result)
            return

        status, payload = final_event
        self._set_processing_state(False)
        if status == "processing_error":
            error_text = str(payload)
            self.status_var.set(error_text)
            return

        if status == "unexpected_error":
            error_text = f"处理失败: {payload}"
            self.status_var.set(error_text)
            return

        result, current_year_only = payload
        self._update_progress(100, "处理完成")
        summary = self._build_summary(result, current_year_only)
        self.status_var.set(summary)

    def _build_summary(self, result: ProcessResult, current_year_only: bool) -> str:
        summary = (
            f"完成: 新增 {result.added_count} 个 ID，保留 {result.preserved_id_count} 个已有 ID，"
            f"有效数据行 {result.content_row_count} 行。\n输出文件: {result.output_path}"
        )
        if current_year_only and result.filtered_row_count is not None:
            summary = (
                f"完成: 新增 {result.added_count} 个 ID，保留 {result.preserved_id_count} 个已有 ID，"
                f"有效数据行 {result.content_row_count} 行，保留本年度数据 {result.filtered_row_count} 行。"
                f"\n输出文件: {result.output_path}"
            )
        return summary

    def _set_processing_state(self, processing: bool, message: str | None = None) -> None:
        self.processing = processing
        if message is not None:
            self.status_var.set(message)

        button_state = "disabled" if processing else "normal"
        combo_state = "disabled" if processing else self._current_sheet_selector_state()

        if self.pick_button is not None:
            self.pick_button.configure(state=button_state)
        if self.export_button is not None:
            self.export_button.configure(state=button_state)
        if self.year_export_button is not None:
            self.year_export_button.configure(state=button_state)
        if self.sheet_combo is not None:
            self.sheet_combo.configure(state=combo_state)
        for button in self.format_buttons:
            button.configure(state=button_state)

        if self.progress_bar is not None:
            if processing:
                self.progress_bar.master.grid()
                self.progress_bar.configure(value=0)
                self.progress_text_var.set("0%")
            else:
                self.progress_bar.configure(value=min(100, float(self.progress_bar.cget("value"))))
                if not self.progress_text_var.get():
                    self.progress_text_var.set("100%")

    def _update_progress(self, percent: int, message: str) -> None:
        if self.progress_bar is not None:
            if not self.progress_bar.master.winfo_ismapped():
                self.progress_bar.master.grid()
            self.progress_bar.configure(value=percent)
        self.progress_text_var.set(f"{percent}%")
        self.status_var.set(message)

    def _current_sheet_selector_state(self) -> str:
        if self.sheet_combo is None:
            return "disabled"
        values = self.sheet_combo.cget("values")
        return "readonly" if values else "disabled"


def main() -> None:
    root = TkinterDnD.Tk() if HAS_DND else tk.Tk()
    IdFillApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
